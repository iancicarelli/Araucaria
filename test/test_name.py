import unittest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from utils.names_processor import NamesProcessor


class TestNamesProcessor(unittest.TestCase):

    def setUp(self):
        # Cargar los archivos Excel desde las rutas proporcionadas
        data_path = r"C:\Users\HP\Desktop\DataFormato.xlsx"
        file_path = r"C:\Users\HP\Desktop\formatoEntrada.xlsx"
        
        # Abrir las hojas de los archivos Excel
        # Data
        data_workbook = load_workbook(data_path)
        self.data_sheet = data_workbook.active
        # file 
        formato_workbook = load_workbook(file_path)
        self.formato_sheet = formato_workbook.active

        # Instancia de NamesProcessor con la hoja de datos
        self.processor = NamesProcessor(self.formato_sheet)
        self.name = NamesProcessor(self.data_sheet)

    def test_get_data_column(self):
        data = self.processor.get_data_column(start_column='AH')  
        self.assertEqual(data, ['40 PANGUIPULLI 2','42 PINTO2','22 PITRUFQUEN',
                                '62 PUREN','11 RECABARREN','47 RUDECINDO ORTEGA','72 SAN JOSE','64 SANTA CRUZ',
                                '38 SCHNEIDER VALDIVIA','37 SIMPSON VALDIVIA','53 TRAIGUEN 2','20 VICTORIA','1 VILCUN'])  #Lista de locales

    def test_extract_names(self):
        values = self.processor.get_data_column(start_column='AH')
        names = self.processor.extract_names(values)
        self.assertEqual(names, ['PANGUIPULLI 2','PINTO2','PITRUFQUEN','PUREN','RECABARREN',
                                'RUDECINDO ORTEGA','SAN JOSE','SANTA CRUZ','SCHNEIDER VALDIVIA','SIMPSON VALDIVIA','TRAIGUEN 2',
                                'VICTORIA','VILCUN'])

    def test_map_names_to_ids(self):
        # Probar que map_names_to_ids devuelve el ID correcto
        values = self.processor.get_data_column(start_column='AH')
        names = self.processor.extract_names(values)
        ids = self.name.map_names_to_ids(names)
        self.assertEqual(ids, 26) 

    

if __name__ == '__main__':
    unittest.main()
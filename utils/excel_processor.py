from openpyxl import load_workbook,Workbook
from models.format import Format
from openpyxl.styles import Font

class ExcelProcessor: 
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None

    def read_excel(self):
        try:
            # Cargar el libro de trabajo
            self.workbook = load_workbook(self.file_path)
            # Seleccionar la hoja activa
            self.sheet = self.workbook.active
        except Exception as e:
            print(f"Error al cargar el archivo Excel: {e}")

    
        
    def write_excel(self, formats, output_path):
    
        try:
            # Crear un nuevo libro de trabajo y una hoja
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Datos Procesados"

            # Agregar las cabeceras a la fila 1
            headers = ["CODIGO", "CANTIDAD", "ID"]
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)  #cabecera en negrita

            # Agregar los datos procesados
            for row_num, format_obj in enumerate(formats, start=2):
                sheet.cell(row=row_num, column=1, value=format_obj.code)
                sheet.cell(row=row_num, column=2, value=format_obj.amount)
                sheet.cell(row=row_num, column=3, value=format_obj.name_id)

            # Guardar el archivo
            workbook.save(output_path)
            print(f"Archivo generado correctamente en: {output_path}")

        except Exception as e:
            print(f"Error al escribir el archivo Excel: {e}")    

    def get_sheet(self):
        return self.sheet
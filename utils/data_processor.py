from openpyxl.utils import column_index_from_string
from models.format import Format

class DataProcessor:
    def __init__(self, sheet,id_sheet):
        self.sheet = sheet
        self.id_sheet = id_sheet


    #ASOCIA EL VALOR TOMADO CON EL CODIGO 
    def process_row(self, id, column_data):
        try:
            formats = []
            verified_data = self.verify_amount(column_data)
            # Obtener todos los códigos (columna B)
            codes = [
                row[1] for row in self.sheet.iter_rows(
                    min_row=3, max_row=self.sheet.max_row, values_only=True
                )
            ]
        
            # Verificar que la cantidad de códigos y datos coincidan
            if len(codes) != len(verified_data):
                print("La cantidad de códigos no coincide con los datos obtenidos.")
                return []

            # Asociar cada código con su respectivo valor
            uxe_data = dict(self.process_uxe())
    
            for code, value in zip(codes, verified_data):
                if code is not None and code in uxe_data and uxe_data[code] != 0:  # Verificar que el código exista y uxe no sea cero
                    division_result = value / uxe_data[code]
                    print(f"Asociando código {code} con valor {division_result}")
                    formats.append(self.create_format(code, division_result, id))

            return formats
        except Exception as e:
            print(f"Error al procesar las filas: {e}")
            return []

    def process_uxe(self):
        try:
            results = []
        
            # Iterar sobre las filas desde la tercera en adelante
            for row in self.sheet.iter_rows(min_row=3, max_row=self.sheet.max_row, values_only=True):
                code = row[1]  # Columna B (índice 1)
                uxe = row[3]  # Columna D (índice 3)
            
                if code is not None and uxe is not None:  # Verificar que no sean None
                    results.append((code, uxe))
        
            return results
        except Exception as e:
            print(f"Error al procesar las filas: {e}")
            return []

    def verify_amount(self, column_data):
        amount_found = []
        for cell in column_data:
            print(f"Valor de la celda: {cell}")  # Para diagnóstico
            # Verificar si el valor es un número (entero o decimal)
            if isinstance(cell, (int, float)):
                amount_found.append(cell)  # Si es numérico, agregarlo
            else:
                # Si no es un número, agregar 0
                amount_found.append(0)
            
        return amount_found



    def create_format(self, code, amount, name_id):
        # Crear un objeto Format con los valores recibidos
        print(f"Creando formato con: Código {code}, Monto {amount}, ID {name_id}")
        return Format(code, amount, name_id)   

    

    #BUSCA EN UNA COLUMNA TODOS LOS VALORES
    def get_data_column(self, row_number=3, start_column='H'):
        if not self.sheet:
            print("No se ha cargado ninguna hoja. Asegúrate de llamar a load_excel primero.")
            return []

        try:
            print(f"Intentando obtener valores desde la fila {row_number}, columna {start_column}")
        
            # Convertir la letra de la columna a índice (1 para 'A', 2 para 'B', etc.)
            start_col_idx = column_index_from_string(start_column)
        
            values = []
            for row in self.sheet.iter_rows(min_row=row_number, max_row=self.sheet.max_row, min_col=start_col_idx, max_col=start_col_idx):
             value = row[0].value  # Extraer el valor de la celda en esa columna
             print(f"Valor en columna ID {start_col_idx}: {value}")
             values.append(value)

            return values

        except Exception as e:
            print(f"Error al obtener valores de la fila {row_number} desde la columna {start_column}: {e}")
        return []

    

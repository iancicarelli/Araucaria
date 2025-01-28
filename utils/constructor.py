from .data_processor import DataProcessor 
from .names_processor import NamesProcessor
from .excel_processor import ExcelProcessor
import string
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


class Constructor:
    def __init__(self, file_path, output_path, data_path, row_processor):
        self.file_path = file_path
        self.output_path = output_path
        self.data_path = data_path
        self.row_processor = row_processor  # Instancia que contiene el método process_row
        
        # Instancias de procesadores
        self.id_processor = ExcelProcessor(self.file_path)
        self.name_processor = ExcelProcessor(self.data_path)
        self.names_processor = None
        self.id_names_processor = None
        self.data_processor = None  
    
    def initialize_processors(self):
        # Leer y cargar datos del archivo principal
        self.id_processor.read_excel()
        id_sheet = self.id_processor.get_sheet()
        self.names_processor = NamesProcessor(id_sheet)

        # Leer y cargar datos del archivo de nombres
        self.name_processor.read_excel()
        data_sheet = self.name_processor.get_sheet()
        self.id_names_processor = NamesProcessor(data_sheet)
        self.data_processor = DataProcessor(id_sheet, data_sheet)


    def get_last_column_with_value(self):
        # Asumimos que estamos en la fila 2
        row = self.id_processor.get_sheet()[2]
        
        # Recorremos las celdas de la fila 2 de derecha a izquierda
        for cell in reversed(row):
            if cell.value is not None:  # Si la celda tiene un valor
                return get_column_letter(cell.column)  # Retorna la letra de la columna

        return None

    def generate_column_letters_up_to(self, last_column_letter,start_column_letter="H"):
        if last_column_letter is None:
            raise ValueError("No se proporcionó una letra válida para la última columna.")
        
        # Convertimos la letra de columna a índice numérico
        start_column_index = column_index_from_string(start_column_letter)
        last_column_index = column_index_from_string(last_column_letter)
        
        # Generamos las letras de columna desde la columna de inicio hasta la última columna indicada
        return [get_column_letter(i) for i in range(start_column_index, last_column_index + 1)]

    def iterate_columns(self):
        if not self.names_processor or not self.id_names_processor:
            raise ValueError("Procesadores no inicializados. Llama a 'initialize_processors' primero.")

        last_column_letter = self.get_last_column_with_value()  # Obtenemos la última columna con valor
        if last_column_letter is None:
            raise ValueError("No se encontró ninguna columna con valor en la fila 2.")    
        
        column_letters = self.generate_column_letters_up_to(last_column_letter)
        
        all_formats = []

        for col_letter in column_letters:
           
            column_data2= self.names_processor.get_data_column(start_column=col_letter)
            names = self.names_processor.extract_names(column_data2)
            ids = self.id_names_processor.map_names_to_ids(names)
            print("IDE", ids)    
            column_data = self.data_processor.get_data_column(start_column=col_letter)
            print(f"Datos obtenidos de la columna {col_letter}: {column_data}: {column_data2}")
            
            # Delegar el procesamiento de filas al método process_row existente
            formats = self.row_processor.process_row(ids, column_data)
            all_formats.extend(formats)
            print(f"Datos procesados para la columna {col_letter}")
        
        return all_formats



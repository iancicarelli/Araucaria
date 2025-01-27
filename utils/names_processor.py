from openpyxl.utils import column_index_from_string
from models.format import Format

class NamesProcessor:

    def __init__(self, sheet):
        self.sheet = sheet

    def get_data_column(self, row_number=2, start_column='H'):
        if not self.sheet:
            print("No se ha cargado ninguna hoja. Asegúrate de llamar a load_excel primero.")
            return []

        try:
            # Convertir la columna inicial a índice (1 para A, 8 para H, etc.)
            start_col_idx = column_index_from_string(start_column)
            max_col_idx = self.sheet.max_column
            # Extraer los valores desde la columna inicial hasta el final
            values = []
            for col_idx in range(start_col_idx, max_col_idx + 1):
                value = self.sheet.cell(row=row_number, column=col_idx).value
                values.append(value)

            return values

        except Exception as e:
            print(f"Error al obtener valores de la fila {row_number} desde la columna {start_column}: {e}")
            return []

    def extract_names(self, values):
        try:
            names = []
            for value in values:
                if isinstance(value, str):
                    parts = value.split(" ", 1)  # Dividir en 2 partes máximo
                    if len(parts) > 1:
                        names.append(parts[1])  # Tomar la segunda parte (el nombre)
            return names

        except Exception as e:
            print(f"Error al extraer los nombres: {e}")
            return []

    def map_names_to_ids(self, names):
        if not self.sheet:
            print("No se ha cargado ninguna hoja. Asegúrate de cargar un archivo Excel.")
            return None

        try:
            for row in self.sheet.iter_rows(min_row=2,max_row=self.sheet.max_row,values_only=True):
                excel_id = row[0]
                excel_name = row[1]

                if excel_name and excel_id:
                    for name in names:
                        if name.strip().upper() == excel_name.strip().upper():
                            return excel_id
                        
            return None
                           
        except Exception as e:
            print(f"Error al mapear nombres a IDs: {e}")
            return None